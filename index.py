import requests
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet('삼성전자')
ws['A1']


codes = ['005930', '000660', '035420']

for code in codes:
    response = requests.get(f'https://finance.naver.com/item/sise.naver?code={code}')
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price =soup.select_one('#_nowVal').text
    price = price.replace(',','')
    print(price)